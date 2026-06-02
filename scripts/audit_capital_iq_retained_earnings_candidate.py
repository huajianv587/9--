#!/usr/bin/env python3
"""Audit Capital IQ retained-earnings candidate exports.

Retained earnings is needed for Altman-style balance-sheet construction. This
audit verifies the field key, target fiscal-year parameters, universe boundary,
and duplicate/leakage columns before the workbook is allowed into cleaning.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!"}
TARGET_FISCAL_YEARS = list(range(2014, 2025))
RETAINED_EARNINGS_KEYS = {
    "IQ_RETAINED_EARNINGS",
    "SNL_RETAINED_EARN",
    "SNL_RETAINED_EARNINGS",
    "SNL_RETAINED_EARN_OTHER_COMMON_EQUITY",
    "SNL_RETAINED_EARN_RESV",
    "PCD_RETAINED_EARNINGS",
    "INS_RETAINED_EARNINGS",
    "INS_CI_RETAINED_EARNINGS",
}
PREFERRED_RETAINED_EARNINGS_KEYS = {"IQ_RETAINED_EARNINGS"}
LEAKAGE_PARAMS = {
    "Latest Fiscal Year",
    "Latest Fiscal Quarter",
    "Year-to-Date",
    "Last Twelve Months",
    "Latest Half-Year",
    "FY2025",
}


def usable_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in NA_STRINGS:
            return None
        stripped = stripped.replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def nonmissing(value: Any) -> bool:
    return value is not None and not (
        isinstance(value, str) and value.strip().upper() in NA_STRINGS
    )


def load_workbook_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError(f"`Sheet1` not found in {workbook}")
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [str(value or "").strip() for value in rows[2]]
    keys = [str(value or "").strip() for value in rows[3]]
    params = [str(value or "").strip() for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def universe_summary(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    ids = []
    exchanges = []
    company_types = []
    for row in rows:
        try:
            ids.append(int(row[1]))
        except (TypeError, ValueError):
            pass
        if len(row) > 2 and row[2] not in (None, ""):
            exchanges.append(str(row[2]).strip())
        if len(row) > 3 and row[3] not in (None, ""):
            company_types.append(str(row[3]).strip())
    return {
        "data_rows": len(rows),
        "unique_ids": len(set(ids)),
        "duplicate_id_rows": len(ids) - len(set(ids)),
        "exchange_counts": Counter(exchanges),
        "company_type_counts": Counter(company_types),
    }


def field_coverage(
    headers: list[str],
    keys: list[str],
    params: list[str],
    rows: list[tuple[Any, ...]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    grouped_cols: dict[tuple[str, str], list[int]] = defaultdict(list)
    for idx, key in enumerate(keys):
        if key in RETAINED_EARNINGS_KEYS:
            grouped_cols[(key, params[idx])].append(idx)

    records = []
    for (key, param), col_idxs in sorted(grouped_cols.items(), key=lambda item: (item[0][0], item[0][1])):
        numeric_by_col = []
        nonmissing_by_col = []
        for col_idx in col_idxs:
            values = [row[col_idx] if col_idx < len(row) else None for row in rows]
            numeric_by_col.append(sum(usable_number(value) is not None for value in values))
            nonmissing_by_col.append(sum(nonmissing(value) for value in values))
        records.append(
            {
                "field_key": key,
                "header": headers[col_idxs[0]] if col_idxs else "",
                "parameter": param or "Blank",
                "column_numbers": [idx + 1 for idx in col_idxs],
                "duplicate_count": len(col_idxs),
                "nonmissing_max": max(nonmissing_by_col) if nonmissing_by_col else 0,
                "numeric_max": max(numeric_by_col) if numeric_by_col else 0,
                "nonmissing_by_col": nonmissing_by_col,
                "numeric_by_col": numeric_by_col,
            }
        )

    available_years = {
        int(match.group(1))
        for key, param in grouped_cols
        if (match := re.fullmatch(r"FY(\d{4})", param))
        and int(match.group(1)) in TARGET_FISCAL_YEARS
    }
    available_keys = sorted({key for key, _ in grouped_cols})
    leakage_cols = {
        f"{key}:{param}": [idx + 1 for idx in cols]
        for (key, param), cols in grouped_cols.items()
        if param in LEAKAGE_PARAMS
    }
    out_of_window_params = sorted(
        {
            param
            for _, param in grouped_cols
            if (match := re.fullmatch(r"FY(\d{4})", param))
            and int(match.group(1)) not in TARGET_FISCAL_YEARS
        }
    )
    duplicate_target_params = sorted(
        f"{key}:{param}"
        for (key, param), cols in grouped_cols.items()
        if len(cols) > 1 and re.fullmatch(r"FY\d{4}", param)
    )
    target_records = [
        record
        for record in records
        if re.fullmatch(r"FY\d{4}", record["parameter"])
        and int(record["parameter"][2:]) in TARGET_FISCAL_YEARS
    ]
    return records, {
        "available_keys": available_keys,
        "available_years": sorted(available_years),
        "missing_years": sorted(set(TARGET_FISCAL_YEARS) - available_years),
        "leakage_cols": leakage_cols,
        "out_of_window_params": out_of_window_params,
        "duplicate_target_params": duplicate_target_params,
        "has_preferred_key": any(key in PREFERRED_RETAINED_EARNINGS_KEYS for key in available_keys),
        "has_candidate_key": any(key in RETAINED_EARNINGS_KEYS for key in available_keys),
        "target_numeric_total": sum(record["numeric_max"] for record in target_records),
    }


def render_report(
    workbook: Path,
    records: list[dict[str, Any]],
    meta: dict[str, Any],
    universe: dict[str, Any],
    expected_exchanges: list[str],
) -> str:
    target_records = [
        record
        for record in records
        if re.fullmatch(r"FY\d{4}", record["parameter"])
        and int(record["parameter"][2:]) in TARGET_FISCAL_YEARS
    ]

    lines = [
        "# Capital IQ Retained Earnings Candidate Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: `{workbook}`",
        "",
        "## Universe",
        "",
        f"- Data rows with Entity ID: {universe['data_rows']}",
        f"- Unique Entity IDs: {universe['unique_ids']}",
        f"- Duplicate Entity-ID rows: {universe['duplicate_id_rows']}",
        f"- Exchange counts: {dict(universe['exchange_counts'])}",
        f"- Company-type counts: {dict(universe['company_type_counts'])}",
        "",
        "## Field Source And Parameter Coverage",
        "",
        f"- Retained-earnings field keys detected: {meta['available_keys']}",
        f"- Preferred CIQ retained-earnings key detected: {meta['has_preferred_key']}",
        f"- Target FY years present: {meta['available_years']}",
        f"- Missing target FY years: {meta['missing_years']}",
        f"- Duplicate target-year retained-earnings parameters: {meta['duplicate_target_params'] or 'None'}",
        f"- Leakage/current columns to exclude in cleaning: {meta['leakage_cols'] or 'None'}",
        f"- Out-of-window FY parameters to exclude in cleaning: {meta['out_of_window_params'] or 'None'}",
        "",
        "## Target-Year Numeric Coverage",
        "",
        "| Field key | Parameter | Columns | Duplicate count | Max non-missing | Max numeric | Numeric by column |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for record in sorted(target_records, key=lambda row: row["parameter"]):
        lines.append(
            f"| `{record['field_key']}` | `{record['parameter']}` | "
            f"{record['column_numbers']} | {record['duplicate_count']} | "
            f"{record['nonmissing_max']} | {record['numeric_max']} | {record['numeric_by_col']} |"
        )

    hard_failures = []
    warnings = []
    if universe["unique_ids"] == 0:
        hard_failures.append("No unique Entity IDs detected.")
    if expected_exchanges:
        exchange_counts = universe["exchange_counts"]
        unexpected = sorted(set(exchange_counts) - set(expected_exchanges))
        missing_expected = [
            exchange for exchange in expected_exchanges if exchange_counts.get(exchange, 0) == 0
        ]
        if missing_expected:
            hard_failures.append(f"Expected exchange(s) not detected: {missing_expected}.")
        if unexpected:
            hard_failures.append(f"Unexpected exchange(s) detected: {unexpected}.")
        expected_total = sum(exchange_counts.get(exchange, 0) for exchange in expected_exchanges)
        if expected_total != universe["data_rows"]:
            warnings.append(
                "Expected exchange rows do not fully reconcile to data rows: "
                f"{dict(exchange_counts)}."
            )
    if not meta["has_candidate_key"]:
        hard_failures.append("No recognized retained-earnings candidate key detected.")
    if meta["missing_years"]:
        hard_failures.append(f"Missing target retained-earnings FY years: {meta['missing_years']}.")
    if not meta["has_preferred_key"]:
        warnings.append(
            "A retained-earnings candidate exists, but the preferred `IQ_RETAINED_EARNINGS` key "
            "was not detected; retain this warning until source-family choice is reconciled."
        )
    if meta["target_numeric_total"] == 0:
        hard_failures.append("Target retained-earnings FY columns have zero numeric coverage.")
    if meta["duplicate_target_params"]:
        warnings.append(
            "Duplicate target-year retained-earnings columns are present; cleaning must deduplicate "
            "by field key and period before merging."
        )
    if meta["leakage_cols"]:
        warnings.append("Latest/current/FY2025 columns are present and must be excluded before modeling.")
    if meta["out_of_window_params"]:
        warnings.append("FY parameters outside 2014-2024 are present and must be excluded before modeling.")

    lines.extend(["", "## Audit Decision", ""])
    if hard_failures:
        lines.append("Status: FAIL")
        lines.extend([f"- {item}" for item in hard_failures])
    else:
        lines.append("Status: PASS_WITH_NOTES")
        lines.append("- Target-year retained-earnings columns are present and have auditable numeric coverage.")
        lines.extend([f"- {item}" for item in warnings])
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--expected-exchange",
        help="Expected exchange name or comma-separated exchange names, e.g. SGX,Catalist.",
    )
    parser.add_argument("--out-md", type=Path)
    args = parser.parse_args()

    headers, keys, params, rows = load_workbook_rows(args.workbook)
    records, meta = field_coverage(headers, keys, params, rows)
    expected_exchanges = [
        item.strip()
        for item in (args.expected_exchange or "").split(",")
        if item.strip()
    ]
    report = render_report(
        workbook=args.workbook,
        records=records,
        meta=meta,
        universe=universe_summary(rows),
        expected_exchanges=expected_exchanges,
    )
    if args.out_md:
        args.out_md.parent.mkdir(parents=True, exist_ok=True)
        args.out_md.write_text(report, encoding="utf-8")
        print(f"Wrote {args.out_md}")
    else:
        print(report)


if __name__ == "__main__":
    main()
