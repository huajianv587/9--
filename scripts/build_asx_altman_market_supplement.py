#!/usr/bin/env python3
"""Build a clean ASX firm-year supplement from a Capital IQ workbook."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
DEFAULT_WORKBOOK = (
    ROOT
    / "data/raw/capital_iq/capital_iq_asx_altman_liquidity_marketcap_2014_2024_attempt_20260602.xlsx"
)
DEFAULT_OUT = ROOT / "data/processed/capital_iq_asx_altman_market_supplement_20260602.csv"
DEFAULT_AUDIT = ROOT / "outputs/asx_altman_market_supplement_merge_audit_20260602.md"

NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!"}

MARKET_CAP_DATE_BY_YEAR = {
    2024: "12/31/2024",
    2023: "12/29/2023",
    2022: "12/30/2022",
    2021: "12/31/2021",
    2020: "12/31/2020",
    2019: "12/31/2019",
    2018: "12/31/2018",
    2017: "12/29/2017",
    2016: "12/30/2016",
    2015: "12/31/2015",
    # 2014 is intentionally omitted until a valid historical market-cap export exists.
}
TARGET_FISCAL_YEARS = list(range(2014, 2025))


def usable_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in NA_STRINGS:
            return None
        stripped = stripped.replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_capital_iq_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [str(value or "") for value in rows[2]]
    keys = [str(value or "") for value in rows[3]]
    params = [str(value or "") for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def build_supplement(workbook: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    headers, keys, params, rows = load_capital_iq_rows(workbook)

    current_assets_cols: dict[int, int] = {}
    current_liab_cols: dict[int, int] = {}
    market_cap_cols: dict[int, tuple[int, str]] = {}
    duplicate_market_params: dict[str, list[int]] = defaultdict(list)
    invalid_market_params: dict[str, int] = {}

    for idx, key in enumerate(keys):
        param = params[idx]
        if key == "SP_CURRENT_ASSETS":
            match = re.fullmatch(r"FY(\d{4})", param)
            if match:
                current_assets_cols[int(match.group(1))] = idx
        elif key == "SP_CURRENT_LIAB":
            match = re.fullmatch(r"FY(\d{4})", param)
            if match:
                current_liab_cols[int(match.group(1))] = idx
        elif key == "SP_MARKETCAP":
            duplicate_market_params[param].append(idx + 1)
            for year, required_date in MARKET_CAP_DATE_BY_YEAR.items():
                if param == required_date and year not in market_cap_cols:
                    market_cap_cols[year] = (idx, param)
            if param in {"Current", "12/31/2023", "12/31/2022"}:
                invalid_market_params[param] = idx + 1

    records: list[dict[str, Any]] = []
    years = TARGET_FISCAL_YEARS
    for row in rows:
        company_name = row[0]
        company_id = int(row[1])
        exchange = row[3] if len(row) > 3 else None
        for year in years:
            current_assets = (
                usable_number(row[current_assets_cols[year]])
                if year in current_assets_cols and current_assets_cols[year] < len(row)
                else None
            )
            current_liabilities = (
                usable_number(row[current_liab_cols[year]])
                if year in current_liab_cols and current_liab_cols[year] < len(row)
                else None
            )
            market_cap = None
            market_cap_date = None
            if year in market_cap_cols:
                col_idx, date = market_cap_cols[year]
                market_cap = usable_number(row[col_idx]) if col_idx < len(row) else None
                market_cap_date = date
            records.append(
                {
                    "company_id": company_id,
                    "company_name": company_name,
                    "exchange": exchange,
                    "fiscal_year": year,
                    "current_assets_usd_000": current_assets,
                    "current_liabilities_usd_000": current_liabilities,
                    "market_cap_usd_m": market_cap,
                    "market_cap_date": market_cap_date,
                }
            )

    supplement = pd.DataFrame.from_records(records)
    duplicate_notes = {
        param: cols
        for param, cols in duplicate_market_params.items()
        if param and len(cols) > 1
    }
    meta = {
        "workbook_rows": len(rows),
        "current_assets_years": sorted(current_assets_cols),
        "current_liab_years": sorted(current_liab_cols),
        "market_cap_years": sorted(market_cap_cols),
        "missing_market_cap_years": sorted(set(TARGET_FISCAL_YEARS) - set(market_cap_cols)),
        "duplicate_market_params": duplicate_notes,
        "invalid_market_params": invalid_market_params,
    }
    return supplement, meta


def render_audit(supplement: pd.DataFrame, panel: pd.DataFrame, meta: dict[str, Any], workbook: Path, out_csv: Path) -> str:
    asx_panel = panel[panel["exchange"].astype(str).str.upper().eq("ASX")].copy()
    asx_panel["company_id"] = pd.to_numeric(asx_panel["company_id"], errors="coerce").astype("Int64")
    asx_panel["fiscal_year"] = pd.to_numeric(asx_panel["fiscal_year"], errors="coerce").astype("Int64")
    base = asx_panel[["company_id", "fiscal_year"]].drop_duplicates()
    merged = base.merge(supplement, on=["company_id", "fiscal_year"], how="left")

    coverage = []
    for year, group in merged.groupby("fiscal_year", dropna=True):
        coverage.append(
            {
                "fiscal_year": int(year),
                "panel_rows": int(len(group)),
                "current_assets": int(group["current_assets_usd_000"].notna().sum()),
                "current_liabilities": int(group["current_liabilities_usd_000"].notna().sum()),
                "market_cap": int(group["market_cap_usd_m"].notna().sum()),
                "market_cap_date": "; ".join(sorted(set(group["market_cap_date"].dropna().astype(str)))),
            }
        )
    coverage_df = pd.DataFrame(coverage).sort_values("fiscal_year")

    lines = [
        "# ASX Altman/Market Supplement Merge Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Source workbook: `{workbook}`",
        f"Output CSV: `{out_csv}`",
        "",
        "## Workbook Parse",
        "",
        f"- Workbook company rows: {meta['workbook_rows']}",
        f"- Current-assets FY years parsed: {meta['current_assets_years']}",
        f"- Current-liabilities FY years parsed: {meta['current_liab_years']}",
        f"- Market-cap years parsed: {meta['market_cap_years']}",
        f"- Missing market-cap years: {meta['missing_market_cap_years']}",
        f"- Duplicate market-cap parameters: {meta['duplicate_market_params'] or 'None'}",
        f"- Invalid market-cap parameters excluded: {meta['invalid_market_params'] or 'None'}",
        "",
        "## ASX Panel Merge Coverage",
        "",
        "| Fiscal year | ASX panel rows | Current assets | Current liabilities | Historical market cap | Market-cap date |",
        "|---:|---:|---:|---:|---:|---|",
    ]
    for row in coverage_df.to_dict("records"):
        lines.append(
            f"| {row['fiscal_year']} | {row['panel_rows']} | {row['current_assets']} | "
            f"{row['current_liabilities']} | {row['market_cap']} | {row['market_cap_date']} |"
        )

    lines.extend(
        [
            "",
            "## Cleaning Rules Applied",
            "",
            "- Historical market capitalization uses explicit date columns only.",
            "- `Current` market capitalization is excluded to avoid look-ahead leakage.",
            "- `12/31/2023` and `12/31/2022` are excluded because the audited export had zero numeric observations.",
            "- Duplicate historical-date columns are reduced by retaining the first occurrence.",
            "- Market cap remains in USD millions; current assets and current liabilities remain in USD thousands.",
            "- This supplement does not claim a complete Altman Z-score because retained earnings and full total-liabilities supplements are still unresolved.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--panel", type=Path, default=PANEL)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--audit-md", type=Path, default=DEFAULT_AUDIT)
    args = parser.parse_args()

    supplement, meta = build_supplement(args.workbook)
    panel = pd.read_csv(args.panel, low_memory=False)
    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    args.audit_md.parent.mkdir(parents=True, exist_ok=True)
    supplement.to_csv(args.out_csv, index=False)
    args.audit_md.write_text(render_audit(supplement, panel, meta, args.workbook, args.out_csv), encoding="utf-8")
    print(f"Wrote {args.out_csv}")
    print(f"Wrote {args.audit_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
