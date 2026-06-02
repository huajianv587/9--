#!/usr/bin/env python3
"""Build a clean Catalist pilot supplement from a Capital IQ workbook."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
DEFAULT_WORKBOOK = (
    ROOT
    / "data/raw/capital_iq/capital_iq_catalist_altman_liquidity_marketcap_pilot_20260602.xlsx"
)
DEFAULT_OUT = ROOT / "data/processed/capital_iq_catalist_altman_market_pilot_20260602.csv"
DEFAULT_AUDIT = ROOT / "outputs/catalist_altman_market_pilot_merge_audit_20260602.md"

NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!"}
TARGET_FISCAL_YEARS = list(range(2014, 2025))
MARKET_CAP_DATE_BY_YEAR = {
    2024: "12/31/2024",
}


def usable_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in NA_STRINGS:
            return None
        stripped = stripped.replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_capital_iq_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [str(value or "") for value in rows[2]]
    keys = [str(value or "") for value in rows[3]]
    params = [str(value or "") for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def build_supplement(workbook: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    _headers, keys, params, rows = load_capital_iq_rows(workbook)

    current_assets_cols: dict[int, int] = {}
    current_liab_cols: dict[int, int] = {}
    market_cap_cols: dict[int, tuple[int, str]] = {}
    duplicate_market_params: dict[str, list[int]] = defaultdict(list)
    ignored_market_params: dict[str, list[int]] = defaultdict(list)

    for idx, key in enumerate(keys):
        param = params[idx]
        if key == "SP_CURRENT_ASSETS":
            match = re.fullmatch(r"FY(\d{4})", param)
            if match and int(match.group(1)) in TARGET_FISCAL_YEARS:
                current_assets_cols[int(match.group(1))] = idx
        elif key == "SP_CURRENT_LIAB":
            match = re.fullmatch(r"FY(\d{4})", param)
            if match and int(match.group(1)) in TARGET_FISCAL_YEARS:
                current_liab_cols[int(match.group(1))] = idx
        elif key == "SP_MARKETCAP":
            duplicate_market_params[param].append(idx + 1)
            matched_year = None
            for year, required_date in MARKET_CAP_DATE_BY_YEAR.items():
                if param == required_date:
                    matched_year = year
                    if year not in market_cap_cols:
                        market_cap_cols[year] = (idx, param)
            if matched_year is None:
                ignored_market_params[param].append(idx + 1)

    records: list[dict[str, Any]] = []
    for row in rows:
        company_name = row[0]
        company_id = int(row[1])
        exchange = row[2] if len(row) > 2 else None
        for year in TARGET_FISCAL_YEARS:
            current_assets = (
                usable_number(row[current_assets_cols[year]])
                if year in current_assets_cols and current_assets_cols[year] < len(row)
                else None
            )
            current_liabilities = (
                usable_number(row[current_liab_cols[year]])
                if year in current_liab_cols and current_liab_cols[year] < len(row)
                else None
            )
            market_cap = None
            market_cap_date = None
            if year in market_cap_cols:
                col_idx, date = market_cap_cols[year]
                market_cap = usable_number(row[col_idx]) if col_idx < len(row) else None
                market_cap_date = date
            records.append(
                {
                    "company_id": company_id,
                    "company_name": company_name,
                    "exchange_current": exchange,
                    "fiscal_year": year,
                    "current_assets_usd_000": current_assets,
                    "current_liabilities_usd_000": current_liabilities,
                    "market_cap_usd_m": market_cap,
                    "market_cap_date": market_cap_date,
                }
            )

    supplement = pd.DataFrame.from_records(records)
    duplicate_notes = {
        param: cols
        for param, cols in duplicate_market_params.items()
        if param and len(cols) > 1
    }
    meta = {
        "workbook_rows": len(rows),
        "current_assets_years": sorted(current_assets_cols),
        "current_liab_years": sorted(current_liab_cols),
        "market_cap_years": sorted(market_cap_cols),
        "missing_market_cap_years": sorted(set(TARGET_FISCAL_YEARS) - set(market_cap_cols)),
        "duplicate_market_params": duplicate_notes,
        "ignored_market_params": {k: v for k, v in ignored_market_params.items() if k},
    }
    return supplement, meta


def render_audit(supplement: pd.DataFrame, panel: pd.DataFrame, meta: dict[str, Any], workbook: Path, out_csv: Path) -> str:
    sg_mask = panel["country"].astype(str).str.casefold().eq("singapore")
    if "exchange" in panel.columns:
        sg_mask = sg_mask | panel["exchange"].astype(str).str.upper().eq("SINGAPORE")
    sg_panel = panel[sg_mask].copy()
    sg_panel["company_id"] = pd.to_numeric(sg_panel["company_id"], errors="coerce").astype("Int64")
    sg_panel["fiscal_year"] = pd.to_numeric(sg_panel["fiscal_year"], errors="coerce").astype("Int64")
    base = sg_panel[["company_id", "fiscal_year"]].drop_duplicates()
    merged = base.merge(supplement, on=["company_id", "fiscal_year"], how="left")

    coverage = []
    for year, group in merged.groupby("fiscal_year", dropna=True):
        coverage.append(
            {
                "fiscal_year": int(year),
                "panel_rows": int(len(group)),
                "current_assets": int(group["current_assets_usd_000"].notna().sum()),
                "current_liabilities": int(group["current_liabilities_usd_000"].notna().sum()),
                "market_cap": int(group["market_cap_usd_m"].notna().sum()),
                "market_cap_date": "; ".join(sorted(set(group["market_cap_date"].dropna().astype(str)))),
            }
        )
    coverage_df = pd.DataFrame(coverage).sort_values("fiscal_year")

    lines = [
        "# Catalist Altman/Market Pilot Merge Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Source workbook: `{workbook}`",
        f"Output CSV: `{out_csv}`",
        "",
        "## Workbook Parse",
        "",
        f"- Workbook company rows: {meta['workbook_rows']}",
        f"- Current-assets FY years parsed: {meta['current_assets_years']}",
        f"- Current-liabilities FY years parsed: {meta['current_liab_years']}",
        f"- Market-cap years parsed: {meta['market_cap_years']}",
        f"- Missing market-cap years: {meta['missing_market_cap_years']}",
        f"- Duplicate market-cap parameters: {meta['duplicate_market_params'] or 'None'}",
        f"- Ignored market-cap parameters: {meta['ignored_market_params'] or 'None'}",
        "",
        "## Singapore Panel Merge Coverage",
        "",
        "| Fiscal year | Singapore panel rows | Current assets | Current liabilities | Historical market cap | Market-cap date |",
        "|---:|---:|---:|---:|---:|---|",
    ]
    for row in coverage_df.to_dict("records"):
        lines.append(
            f"| {row['fiscal_year']} | {row['panel_rows']} | {row['current_assets']} | "
            f"{row['current_liabilities']} | {row['market_cap']} | {row['market_cap_date']} |"
        )

    lines.extend(
        [
            "",
            "## Cleaning Rules Applied",
            "",
            "- Only explicit fiscal-year current-assets and current-liabilities columns are retained.",
            "- Market capitalization is retained only for explicit historical date columns listed in the script.",
            "- Duplicate market-cap date columns are reduced by retaining the first occurrence.",
            "- Market cap remains in USD millions; current assets and current liabilities remain in USD thousands.",
            "- This is a Catalist pilot, not a complete Singapore supplement.",
            "- This supplement does not claim a complete Altman Z-score because retained earnings, full total liabilities, SGX mainboard coverage, and event dates are still unresolved.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--panel", type=Path, default=PANEL)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--audit-md", type=Path, default=DEFAULT_AUDIT)
    args = parser.parse_args()

    supplement, meta = build_supplement(args.workbook)
    panel = pd.read_csv(args.panel, low_memory=False)
    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    args.audit_md.parent.mkdir(parents=True, exist_ok=True)
    supplement.to_csv(args.out_csv, index=False)
    args.audit_md.write_text(render_audit(supplement, panel, meta, args.workbook, args.out_csv), encoding="utf-8")
    print(f"Wrote {args.out_csv}")
    print(f"Wrote {args.audit_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
