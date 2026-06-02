#!/usr/bin/env python3
"""Audit a supplemental Capital IQ workbook before any model merge."""

from __future__ import annotations

import argparse
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
OUTPUTS = ROOT / "outputs"

ID_ALIASES = {
    "SP_ENTITY_ID",
    "Entity ID",
    "Company ID",
    "Capital IQ Company ID",
    "IQ ID",
    "company_id",
}

DATE_TOKENS = ("date", "as of", "as-of", "fiscal", "fy", "period")
CRITICAL_TOKENS = (
    "current assets",
    "current liabilities",
    "retained earnings",
    "accumulated deficit",
    "market capitalization",
    "market cap",
    "total liabilities",
    "total assets",
    "total revenue",
    "sales",
    "ebit",
    "interest expense",
    "company status",
    "delisting",
    "bankruptcy",
    "default",
    "liquidation",
    "receivership",
    "restructuring",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(rows: list[list[Any]]) -> int | None:
    best_row: int | None = None
    best_score = 0
    for idx, row in enumerate(rows, start=1):
        labels = [norm(value) for value in row]
        nonblank = [label for label in labels if label]
        if len(nonblank) < 2:
            continue
        score = 0
        if any(label in ID_ALIASES for label in nonblank):
            score += 10
        lower = " | ".join(label.lower() for label in nonblank)
        score += sum(token in lower for token in CRITICAL_TOKENS)
        score += min(len(nonblank), 8)
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def clean_header(labels: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    cleaned: list[str] = []
    for i, value in enumerate(labels, start=1):
        label = norm(value) or f"unnamed_{i}"
        if label in seen:
            seen[label] += 1
            label = f"{label}.{seen[label]}"
        else:
            seen[label] = 0
        cleaned.append(label)
    return cleaned


def sheet_summary(path: Path, sheet_name: str, model_ids: set[int]) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[sheet_name]
    preview_rows = [
        list(row)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True)
    ]
    header_row = find_header_row(preview_rows)
    summary: dict[str, Any] = {
        "sheet": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_row": header_row,
        "preview_rows": [
            [norm(value) for value in row[:20]]
            for row in preview_rows[:8]
        ],
        "id_column": "",
        "unique_ids": None,
        "overlap_model_ids": None,
        "missing_model_ids": None,
        "extra_export_ids": None,
        "date_like_columns": [],
        "critical_like_columns": [],
        "top_nonmissing": [],
    }
    if header_row is None:
        return summary

    rows = ws.iter_rows(min_row=header_row, values_only=True)
    header = clean_header(list(next(rows)))
    data = pd.DataFrame(rows, columns=header)
    data = data.dropna(how="all")
    summary["data_rows"] = int(len(data))

    lower_map = {col: col.lower() for col in data.columns}
    date_cols = [
        col
        for col, lower in lower_map.items()
        if any(token in lower for token in DATE_TOKENS)
    ]
    critical_cols = [
        col
        for col, lower in lower_map.items()
        if any(token in lower for token in CRITICAL_TOKENS)
    ]
    summary["date_like_columns"] = date_cols[:40]
    summary["critical_like_columns"] = critical_cols[:80]

    nonmissing = data.notna().sum().sort_values(ascending=False)
    summary["top_nonmissing"] = [
        (str(col), int(count)) for col, count in nonmissing.head(30).items()
    ]

    id_col = next((col for col in data.columns if col in ID_ALIASES), "")
    if not id_col:
        id_col = next((col for col in data.columns if col.lower() in {"sp_entity_id", "entity id", "company id", "company_id"}), "")
    if id_col:
        ids = pd.to_numeric(data[id_col], errors="coerce").dropna().astype(int)
        export_ids = set(ids.tolist())
        summary["id_column"] = id_col
        summary["unique_ids"] = len(export_ids)
        summary["overlap_model_ids"] = len(export_ids & model_ids)
        summary["missing_model_ids"] = len(model_ids - export_ids)
        summary["extra_export_ids"] = len(export_ids - model_ids)

    return summary


def render_report(path: Path, summaries: list[dict[str, Any]]) -> str:
    stat = path.stat()
    lines = [
        "# Capital IQ Supplemental Workbook Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: `{path}`",
        f"Size bytes: {stat.st_size:,}",
        f"Modified: {datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds')}",
        f"SHA-256: `{sha256(path)}`",
        "",
        "## Workbook Summary",
        "",
        "| Sheet | Rows | Columns | Header row | Data rows | ID column | Unique IDs | Model overlap | Missing model IDs | Extra IDs |",
        "|---|---:|---:|---:|---:|---|---:|---:|---:|---:|",
    ]
    for item in summaries:
        lines.append(
            f"| {item['sheet']} | {item['max_row']} | {item['max_column']} | "
            f"{item.get('header_row') or ''} | {item.get('data_rows', '')} | "
            f"{item.get('id_column', '')} | {item.get('unique_ids') or ''} | "
            f"{item.get('overlap_model_ids') or ''} | {item.get('missing_model_ids') or ''} | "
            f"{item.get('extra_export_ids') or ''} |"
        )

    for item in summaries:
        lines.extend(
            [
                "",
                f"## Sheet: {item['sheet']}",
                "",
                "### Date-like columns",
                "",
            ]
        )
        date_cols = item.get("date_like_columns") or []
        lines.extend([f"- `{col}`" for col in date_cols] or ["- None detected"])
        lines.extend(["", "### Critical-like columns", ""])
        critical_cols = item.get("critical_like_columns") or []
        lines.extend([f"- `{col}`" for col in critical_cols] or ["- None detected"])
        lines.extend(["", "### First rows preview", ""])
        preview = item.get("preview_rows") or []
        if preview:
            lines.append("| Row | First non-empty cells |")
            lines.append("|---:|---|")
            for idx, row in enumerate(preview, start=1):
                cells = [cell.replace("\n", " / ") for cell in row if cell]
                lines.append(f"| {idx} | {'; '.join(cells[:12])} |")
        else:
            lines.append("- No preview available")
        lines.extend(["", "### Top non-missing columns", "", "| Column | Non-missing |", "|---|---:|"])
        for col, count in item.get("top_nonmissing") or []:
            lines.append(f"| `{col}` | {count:,} |")

    lines.extend(
        [
            "",
            "## Merge Decision Checklist",
            "",
            "- Do not merge if market-cap/date fields are current or undated.",
            "- Do not merge if the workbook only covers a current operating-company universe.",
            "- Do not merge if critical fiscal-year labels are missing or inconsistent.",
            "- Do not merge if model-ID overlap is materially below the expected research universe without a targeted reconciliation plan.",
            "- Use this report as intake evidence only; model-panel changes require a separate merge script and leakage audit.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="Path to a Capital IQ xlsx workbook")
    parser.add_argument("--out-md", default="", help="Optional output markdown path")
    args = parser.parse_args()

    path = Path(args.workbook).expanduser().resolve()
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")
    if path.suffix.lower() != ".xlsx":
        raise SystemExit("Only .xlsx workbooks are currently supported")

    panel = pd.read_csv(PANEL, usecols=["company_id"], low_memory=False)
    model_ids = set(pd.to_numeric(panel["company_id"], errors="coerce").dropna().astype(int))
    workbook = load_workbook(path, read_only=True, data_only=True)
    summaries = [sheet_summary(path, sheet.title, model_ids) for sheet in workbook.worksheets]
    report = render_report(path, summaries)

    if args.out_md:
        out = (ROOT / args.out_md).resolve() if not Path(args.out_md).is_absolute() else Path(args.out_md)
    else:
        stem = path.stem.replace(" ", "_")
        out = OUTPUTS / f"{stem}_supplemental_workbook_audit.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(report, encoding="utf-8")
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
