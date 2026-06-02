#!/usr/bin/env python3
"""Audit Capital IQ field keys, parameters, and numeric coverage."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"

NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!"}
TARGET_FISCAL_YEARS = list(range(2014, 2025))
EXPECTED_MARKET_CAP_BY_YEAR = {
    2024: "12/31/2024",
    2023: "12/29/2023",
    2022: "12/30/2022",
    2021: "12/31/2021",
    2020: "12/31/2020",
    2019: "12/31/2019",
    2018: "12/31/2018",
    2017: "12/29/2017",
    2016: "12/30/2016",
    2015: "12/31/2015",
    2014: "12/31/2014",
}
CRITICAL_KEYS = {"SP_CURRENT_ASSETS", "SP_CURRENT_LIAB", "SP_MARKETCAP"}


def usable_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in NA_STRINGS:
            return None
        stripped = stripped.replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [str(value or "").strip() for value in rows[2]]
    keys = [str(value or "").strip() for value in rows[3]]
    params = [str(value or "").strip() for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def key_param_coverage(
    headers: list[str],
    keys: list[str],
    params: list[str],
    rows: list[tuple[Any, ...]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    grouped_cols: dict[tuple[str, str], list[int]] = defaultdict(list)
    for idx, key in enumerate(keys):
        if key in CRITICAL_KEYS:
            grouped_cols[(key, params[idx])].append(idx)

    records: list[dict[str, Any]] = []
    for (key, param), col_idxs in sorted(grouped_cols.items(), key=lambda item: (item[0][0], item[0][1])):
        nonmissing_by_col = []
        numeric_by_col = []
        for col_idx in col_idxs:
            values = [row[col_idx] if col_idx < len(row) else None for row in rows]
            nonmissing = sum(
                value is not None
                and not (isinstance(value, str) and value.strip().upper() in NA_STRINGS)
                for value in values
            )
            numeric = sum(usable_number(value) is not None for value in values)
            nonmissing_by_col.append(nonmissing)
            numeric_by_col.append(numeric)
        records.append(
            {
                "field_key": key,
                "header": headers[col_idxs[0]] if col_idxs else "",
                "parameter": param or "Blank",
                "column_numbers": [idx + 1 for idx in col_idxs],
                "duplicate_count": len(col_idxs),
                "nonmissing_max": max(nonmissing_by_col) if nonmissing_by_col else 0,
                "numeric_max": max(numeric_by_col) if numeric_by_col else 0,
                "nonmissing_by_col": nonmissing_by_col,
                "numeric_by_col": numeric_by_col,
            }
        )

    available_current_assets = {
        int(match.group(1))
        for key, param in grouped_cols
        if key == "SP_CURRENT_ASSETS"
        and (match := re.fullmatch(r"FY(\d{4})", param))
        and int(match.group(1)) in TARGET_FISCAL_YEARS
    }
    available_current_liab = {
        int(match.group(1))
        for key, param in grouped_cols
        if key == "SP_CURRENT_LIAB"
        and (match := re.fullmatch(r"FY(\d{4})", param))
        and int(match.group(1)) in TARGET_FISCAL_YEARS
    }
    available_market = {
        year
        for year, date in EXPECTED_MARKET_CAP_BY_YEAR.items()
        if ("SP_MARKETCAP", date) in grouped_cols
    }
    duplicate_params = {
        f"{key}:{param or 'Blank'}": [idx + 1 for idx in col_idxs]
        for (key, param), col_idxs in grouped_cols.items()
        if len(col_idxs) > 1
    }
    market_params = sorted(param for key, param in grouped_cols if key == "SP_MARKETCAP")
    meta = {
        "available_current_assets_years": sorted(available_current_assets),
        "available_current_liab_years": sorted(available_current_liab),
        "available_market_cap_years": sorted(available_market),
        "missing_current_assets_years": sorted(set(TARGET_FISCAL_YEARS) - available_current_assets),
        "missing_current_liab_years": sorted(set(TARGET_FISCAL_YEARS) - available_current_liab),
        "missing_market_cap_years": sorted(set(TARGET_FISCAL_YEARS) - available_market),
        "duplicate_params": duplicate_params,
        "market_params": market_params,
        "has_current_market_cap": ("SP_MARKETCAP", "Current") in grouped_cols,
    }
    return records, meta


def universe_summary(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    ids = []
    exchanges = []
    company_types = []
    for row in rows:
        try:
            ids.append(int(row[1]))
        except (TypeError, ValueError):
            pass
        if len(row) > 2 and row[2] not in (None, ""):
            exchanges.append(str(row[2]).strip())
        if len(row) > 3 and row[3] not in (None, ""):
            company_types.append(str(row[3]).strip())
    return {
        "data_rows": len(rows),
        "unique_ids": len(set(ids)),
        "duplicate_id_rows": len(ids) - len(set(ids)),
        "exchange_counts": Counter(exchanges),
        "company_type_counts": Counter(company_types),
    }


def render_report(
    workbook: Path,
    records: list[dict[str, Any]],
    meta: dict[str, Any],
    universe: dict[str, Any],
) -> str:
    targeted_records = []
    for record in records:
        key = record["field_key"]
        param = record["parameter"]
        if key in {"SP_CURRENT_ASSETS", "SP_CURRENT_LIAB"} and re.fullmatch(r"FY\d{4}", param):
            year = int(param[2:])
            if year in TARGET_FISCAL_YEARS:
                targeted_records.append(record)
        elif key == "SP_MARKETCAP" and param in set(EXPECTED_MARKET_CAP_BY_YEAR.values()) | {"Current"}:
            targeted_records.append(record)

    lines = [
        "# Capital IQ Field-Parameter Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: `{workbook}`",
        "",
        "## Universe",
        "",
        f"- Data rows with Entity ID: {universe['data_rows']}",
        f"- Unique Entity IDs: {universe['unique_ids']}",
        f"- Duplicate Entity-ID rows: {universe['duplicate_id_rows']}",
        f"- Exchange counts: {dict(universe['exchange_counts'])}",
        f"- Company-type counts: {dict(universe['company_type_counts'])}",
        "",
        "## Required Parameter Coverage",
        "",
        f"- Current-assets FY years present: {meta['available_current_assets_years']}",
        f"- Current-liabilities FY years present: {meta['available_current_liab_years']}",
        f"- Market-cap historical years present: {meta['available_market_cap_years']}",
        f"- Missing current-assets FY years: {meta['missing_current_assets_years']}",
        f"- Missing current-liabilities FY years: {meta['missing_current_liab_years']}",
        f"- Missing market-cap historical years: {meta['missing_market_cap_years']}",
        f"- Contains `SP_MARKETCAP` Current column: {meta['has_current_market_cap']}",
        f"- Duplicate critical parameters: {meta['duplicate_params'] or 'None'}",
        "",
        "## Targeted Numeric Coverage",
        "",
        "| Field key | Parameter | Columns | Duplicate count | Max non-missing | Max numeric | Numeric by column |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for record in sorted(targeted_records, key=lambda row: (row["field_key"], row["parameter"])):
        lines.append(
            f"| `{record['field_key']}` | `{record['parameter']}` | "
            f"{record['column_numbers']} | {record['duplicate_count']} | "
            f"{record['nonmissing_max']} | {record['numeric_max']} | {record['numeric_by_col']} |"
        )

    lines.extend(
        [
            "",
            "## Market-Cap Parameter Inventory",
            "",
        ]
    )
    lines.extend([f"- `{param}`" for param in meta["market_params"]] or ["- None"])
    lines.extend(
        [
            "",
            "## Audit Decision",
            "",
        ]
    )
    hard_failures = []
    if universe["unique_ids"] == 0:
        hard_failures.append("No unique Entity IDs detected.")
    if not meta["available_current_assets_years"]:
        hard_failures.append("No target current-assets fiscal-year parameters detected.")
    if not meta["available_current_liab_years"]:
        hard_failures.append("No target current-liabilities fiscal-year parameters detected.")
    if not meta["available_market_cap_years"]:
        hard_failures.append("No expected historical market-cap date parameters detected.")

    if hard_failures:
        lines.append("Status: FAIL")
        lines.extend([f"- {item}" for item in hard_failures])
    else:
        lines.append("Status: PASS_WITH_NOTES")
        lines.append("- Required field families were detected and can be cleaned with explicit leakage controls.")
        if meta["has_current_market_cap"]:
            lines.append("- `SP_MARKETCAP` Current exists and must be excluded from model features.")
        if meta["missing_market_cap_years"]:
            lines.append("- Missing or zero-coverage historical market-cap years must remain documented as gaps.")

    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-md", type=Path, required=True)
    args = parser.parse_args()

    workbook = args.workbook.expanduser().resolve()
    headers, keys, params, rows = load_rows(workbook)
    records, meta = key_param_coverage(headers, keys, params, rows)
    universe = universe_summary(rows)
    report = render_report(workbook, records, meta, universe)
    out_md = args.out_md if args.out_md.is_absolute() else ROOT / args.out_md
    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_md.write_text(report, encoding="utf-8")
    print(f"Wrote {out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
