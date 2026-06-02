#!/usr/bin/env python3
"""Check Chrome-controlled Capital IQ download evidence."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"
PATTERNS = ["SPGlobal_Export*.xlsx", "*Capital*.xlsx", "*.csv", "*.xlsx"]


def parse_local_time(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def recent_downloads(after: datetime) -> list[Path]:
    seen: set[Path] = set()
    rows: list[Path] = []
    for pattern in PATTERNS:
        for path in DOWNLOADS.glob(pattern):
            if path in seen or not path.is_file():
                continue
            seen.add(path)
            if datetime.fromtimestamp(path.stat().st_mtime) > after:
                rows.append(path)
    return sorted(rows, key=lambda path: path.stat().st_mtime, reverse=True)


def summarize_xlsx(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        preview = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
            preview.append([str(value)[:120] if value is not None else None for value in row[:100]])
        sheets.append(
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "preview": preview,
            }
        )
    return {"sheets": sheets}


def describe(path: Path) -> dict[str, object]:
    stat = path.stat()
    row: dict[str, object] = {
        "path": str(path),
        "name": path.name,
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
    }
    if path.suffix.lower() == ".xlsx":
        row.update(summarize_xlsx(path))
    return row


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--after", required=True, help="local timestamp, e.g. '2026-06-01 22:20:00'")
    parser.add_argument("--expect-name", default="", help="optional expected downloaded filename")
    parser.add_argument("--out-json", default="", help="optional JSON output path")
    args = parser.parse_args()

    after = parse_local_time(args.after)
    downloads = [describe(path) for path in recent_downloads(after)]
    result = {
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "after": after.isoformat(timespec="seconds"),
        "expected_name": args.expect_name,
        "expected_found": bool(args.expect_name)
        and any(row.get("name") == args.expect_name for row in downloads),
        "recent_downloads": downloads,
    }
    text = json.dumps(result, indent=2, ensure_ascii=False)
    if args.out_json:
        out = ROOT / args.out_json
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text + "\n", encoding="utf-8")
        print(f"Wrote {out}")
    print(text)
    if args.expect_name and not result["expected_found"]:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
