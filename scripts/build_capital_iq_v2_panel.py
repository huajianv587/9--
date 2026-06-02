#!/usr/bin/env python3
"""Build an audited v2 firm-year panel from Capital IQ supplemental workbooks."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
import hashlib
import re
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data/raw/capital_iq"
BASE_PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
OUT_PANEL = ROOT / "data/processed/ael_apac_firm_year_panel_v2_capital_iq_20260603.csv"
OUT_AUDIT = ROOT / "outputs/ael_apac_firm_year_panel_v2_capital_iq_audit_20260603.md"

TARGET_YEARS = list(range(2014, 2025))
MARKET_CAP_DATE_BY_YEAR = {
    2024: "12/31/2024",
    2023: "12/29/2023",
    2022: "12/30/2022",
    2021: "12/31/2021",
    2020: "12/31/2020",
    2019: "12/31/2019",
    2018: "12/31/2018",
    2017: "12/29/2017",
    2016: "12/30/2016",
    2015: "12/31/2015",
    2014: "12/31/2014",
}

LIQUIDITY_MARKET_WORKBOOKS = [
    RAW / "capital_iq_asx_altman_liquidity_marketcap_2014_2024_full_20260602.xlsx",
    RAW / "capital_iq_sgx_altman_liquidity_marketcap_2014_2024_20260602.xlsx",
    RAW / "capital_iq_catalist_altman_liquidity_marketcap_2014_2024_20260602.xlsx",
]
TOTAL_LIAB_WORKBOOKS = [
    RAW / "capital_iq_asx_total_liabilities_candidate_snl_2014_2024_20260602.xlsx",
    RAW / "capital_iq_sgx_catalist_total_liabilities_candidate_snl_2014_2024_20260602.xlsx",
]
RETAINED_EARNINGS_WORKBOOKS = [
    RAW / "capital_iq_asx_retained_earnings_iq_2014_2024_20260602.xlsx",
    RAW / "capital_iq_sgx_catalist_retained_earnings_iq_2014_2024_20260602.xlsx",
]
STATUS_WORKBOOK = RAW / "capital_iq_aus_sg_public_company_broad_identifier_status_ipo_20260602.xlsx"
EVENT_WORKBOOK = RAW / "capital_iq_key_developments_distress_events_aus_sg_2010_2024_20260602.xlsx"

NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!", "NAN"}
DISTRESS_KEYWORDS = (
    "bankrupt",
    "bankruptcy",
    "delist",
    "delisting",
    "suspend",
    "suspension",
    "default",
    "liquidat",
    "receiver",
    "receivership",
    "administration",
    "restructur",
    "distress",
    "going concern",
    "red flag",
    "winding up",
    "insolv",
    "voluntary administration",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def usable_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() in NA_STRINGS:
            return None
        stripped = stripped.replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    try:
        if pd.isna(value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def usable_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() in NA_STRINGS:
        return None
    return text


def first_number(row: tuple[Any, ...], cols: Iterable[int]) -> float | None:
    for idx in cols:
        if idx < len(row):
            value = usable_number(row[idx])
            if value is not None:
                return value
    return None


def first_text(row: tuple[Any, ...], cols: Iterable[int]) -> str | None:
    for idx in cols:
        if idx < len(row):
            value = usable_text(row[idx])
            if value is not None:
                return value
    return None


def load_capital_iq_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [str(value or "") for value in rows[2]]
    keys = [str(value or "") for value in rows[3]]
    params = [str(value or "") for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def key_param_cols(keys: list[str], params: list[str], target_key: str, target_param: str) -> list[int]:
    return [idx for idx, (key, param) in enumerate(zip(keys, params)) if key == target_key and param == target_param]


def key_cols(keys: list[str], target_key: str) -> list[int]:
    return [idx for idx, key in enumerate(keys) if key == target_key]


def build_liquidity_market(workbooks: list[Path]) -> tuple[pd.DataFrame, dict[str, Any]]:
    records: list[dict[str, Any]] = []
    meta: dict[str, Any] = {"workbooks": [], "duplicate_key_params": defaultdict(int)}
    for workbook in workbooks:
        headers, keys, params, rows = load_capital_iq_rows(workbook)
        entity_name_cols = key_cols(keys, "SP_ENTITY_NAME")
        entity_id_cols = key_cols(keys, "SP_ENTITY_ID")
        exchange_cols = key_cols(keys, "SP_EXCHANGE")
        company_type_cols = key_cols(keys, "SP_COMPANY_TYPE")
        field_counts: dict[str, int] = defaultdict(int)
        duplicate_pairs: dict[str, int] = defaultdict(int)
        for key, param in zip(keys, params):
            if key:
                field_counts[key] += 1
                duplicate_pairs[f"{key}|{param}"] += 1
        for pair, count in duplicate_pairs.items():
            if count > 1:
                meta["duplicate_key_params"][pair] += count

        for row in rows:
            company_id = first_number(row, entity_id_cols)
            if company_id is None:
                continue
            for year in TARGET_YEARS:
                market_cols = key_param_cols(keys, params, "SP_MARKETCAP", MARKET_CAP_DATE_BY_YEAR[year])
                records.append(
                    {
                        "company_id": int(company_id),
                        "fiscal_year": year,
                        "ciq_name_liq": first_text(row, entity_name_cols),
                        "ciq_exchange_liq": first_text(row, exchange_cols),
                        "ciq_company_type_liq": first_text(row, company_type_cols),
                        "current_assets_ciq_usd_000": first_number(row, key_param_cols(keys, params, "SP_CURRENT_ASSETS", f"FY{year}")),
                        "current_liabilities_ciq_usd_000": first_number(row, key_param_cols(keys, params, "SP_CURRENT_LIAB", f"FY{year}")),
                        "market_cap_ciq_usd_m": first_number(row, market_cols),
                        "market_cap_date_ciq": MARKET_CAP_DATE_BY_YEAR[year] if market_cols else None,
                        "liquidity_market_source": workbook.name,
                    }
                )
        meta["workbooks"].append(
            {
                "name": workbook.name,
                "sha256": sha256(workbook),
                "company_rows": len(rows),
                "field_counts": dict(sorted(field_counts.items())),
            }
        )
    df = pd.DataFrame.from_records(records)
    return df.drop_duplicates(["company_id", "fiscal_year"], keep="first"), meta


def build_yearly_field(workbooks: list[Path], key: str, value_name: str, source_name: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    records: list[dict[str, Any]] = []
    meta: dict[str, Any] = {"workbooks": [], "duplicate_key_params": defaultdict(int)}
    for workbook in workbooks:
        _headers, keys, params, rows = load_capital_iq_rows(workbook)
        entity_name_cols = key_cols(keys, "SP_ENTITY_NAME")
        entity_id_cols = key_cols(keys, "SP_ENTITY_ID")
        exchange_cols = key_cols(keys, "SP_EXCHANGE")
        company_type_cols = key_cols(keys, "SP_COMPANY_TYPE")
        duplicate_pairs: dict[str, int] = defaultdict(int)
        for col_key, param in zip(keys, params):
            if col_key:
                duplicate_pairs[f"{col_key}|{param}"] += 1
        for pair, count in duplicate_pairs.items():
            if count > 1:
                meta["duplicate_key_params"][pair] += count
        for row in rows:
            company_id = first_number(row, entity_id_cols)
            if company_id is None:
                continue
            for year in TARGET_YEARS:
                records.append(
                    {
                        "company_id": int(company_id),
                        "fiscal_year": year,
                        f"ciq_name_{source_name}": first_text(row, entity_name_cols),
                        f"ciq_exchange_{source_name}": first_text(row, exchange_cols),
                        f"ciq_company_type_{source_name}": first_text(row, company_type_cols),
                        value_name: first_number(row, key_param_cols(keys, params, key, f"FY{year}")),
                        f"{source_name}_source": workbook.name,
                    }
                )
        meta["workbooks"].append({"name": workbook.name, "sha256": sha256(workbook), "company_rows": len(rows)})
    df = pd.DataFrame.from_records(records)
    return df.drop_duplicates(["company_id", "fiscal_year"], keep="first"), meta


def build_status_universe(workbook: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    _headers, keys, _params, rows = load_capital_iq_rows(workbook)
    records = []
    for row in rows:
        company_id = first_number(row, key_cols(keys, "SP_ENTITY_ID"))
        if company_id is None:
            continue
        records.append(
            {
                "company_id": int(company_id),
                "ciq_status_name": first_text(row, key_cols(keys, "SP_ENTITY_NAME")),
                "ciq_company_type_status": first_text(row, key_cols(keys, "SP_COMPANY_TYPE")),
                "ciq_status": first_text(row, key_cols(keys, "SP_COMPANY_STATUS")),
                "ciq_ipo_date": first_text(row, key_cols(keys, "SP_IPO_DATE")),
                "ciq_geography": first_text(row, key_cols(keys, "SP_GEOGRAPHY")),
                "ciq_exchange_status": first_text(row, key_cols(keys, "SP_EXCHANGE")),
            }
        )
    df = pd.DataFrame.from_records(records).drop_duplicates("company_id", keep="first")
    return df, {"workbook": workbook.name, "sha256": sha256(workbook), "company_rows": len(rows)}


def clean_event_headers(labels: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for idx, value in enumerate(labels, start=1):
        label = str(value or "").strip() or f"unnamed_{idx}"
        if label in seen:
            seen[label] += 1
            label = f"{label}.{seen[label]}"
        else:
            seen[label] = 0
        out.append(label)
    return out


def normalize_company_name(value: Any) -> str:
    text = str(value or "").lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\b(asx|sgx|catalist|ses|sg)\b[:\s]*[a-z0-9.\-]+", " ", text)
    text = re.sub(
        r"\b(limited|ltd|inc|plc|corporation|corp|company|co|group|holdings|holding|pte|pty|trust|stapled securities)\b",
        " ",
        text,
    )
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text)


def strong_name_key(name_key: str) -> bool:
    tokens = name_key.split()
    return len(name_key) >= 8 or len(tokens) >= 2


def candidate_company_names(value: Any) -> list[str]:
    text = str(value or "")
    parts = re.split(r";|\n\n|\n", text)
    return [part.strip() for part in parts if part and part.strip()]


def build_panel_name_map(panel: pd.DataFrame) -> dict[str, int]:
    names = panel[["company_id", "company_name"]].drop_duplicates("company_id").copy()
    names["name_key"] = names["company_name"].map(normalize_company_name)
    counts = names["name_key"].value_counts()
    unique = names.loc[names["name_key"].map(counts).eq(1) & names["name_key"].map(strong_name_key)]
    return dict(zip(unique["name_key"], unique["company_id"]))


def load_events(workbook: Path, panel: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = clean_event_headers(list(rows[2]))
    data = pd.DataFrame(rows[3:], columns=header).dropna(how="all")
    code_like = data.astype(str).apply(lambda row: sum(value.strip().startswith("SPKD_") for value in row), axis=1)
    data = data.loc[code_like < 2].copy()

    id_col = "SPCIQ ID"
    date_col = "Key Development Date\nMM/dd/yyyy"
    text_cols = [
        col
        for col in data.columns
        if any(token in col.lower() for token in ("type", "category", "headline", "description"))
    ]
    model_ids = set(panel["company_id"].dropna().astype(int).tolist())
    name_map = build_panel_name_map(panel)
    data["spciq_numeric_id"] = pd.to_numeric(data[id_col].astype(str).str.extract(r"(\d+)", expand=False), errors="coerce")
    data["event_date"] = pd.to_datetime(data[date_col], errors="coerce")
    data["event_text_blob"] = data[text_cols].fillna("").astype(str).agg(" | ".join, axis=1)
    blob_lower = data["event_text_blob"].str.lower()
    data["distress_keyword_event"] = blob_lower.apply(lambda value: any(keyword in value for keyword in DISTRESS_KEYWORDS))

    mapped_records: list[dict[str, Any]] = []
    for _, row in data.iterrows():
        event_date = row["event_date"]
        if pd.isna(event_date) or not row["distress_keyword_event"]:
            continue
        if not (pd.Timestamp("2010-01-01") <= pd.Timestamp(event_date) <= pd.Timestamp("2024-12-31")):
            continue

        direct_id = row["spciq_numeric_id"]
        matched_ids: set[int] = set()
        match_method = ""
        if not pd.isna(direct_id) and int(direct_id) in model_ids:
            matched_ids.add(int(direct_id))
            match_method = "direct_id"

        for candidate in candidate_company_names(row.get("Company Name", "")):
            key = normalize_company_name(candidate)
            if key in name_map:
                matched_ids.add(int(name_map[key]))
                match_method = "direct_id_or_conservative_name" if match_method else "conservative_name"

        for company_id in matched_ids:
            mapped_records.append(
                {
                    "company_id": company_id,
                    "event_date": pd.Timestamp(event_date),
                    "event_text_blob": row["event_text_blob"],
                    "event_match_method": match_method,
                }
            )

    events = pd.DataFrame.from_records(mapped_records)
    if events.empty:
        events = pd.DataFrame(columns=["company_id", "event_date", "event_text_blob", "event_match_method"])
    else:
        events = events.drop_duplicates(["company_id", "event_date", "event_text_blob"]).copy()
    direct_overlap = set(data["spciq_numeric_id"].dropna().astype(int).tolist()) & model_ids
    meta = {
        "workbook": workbook.name,
        "sha256": sha256(workbook),
        "raw_rows": int(len(data)),
        "parseable_event_rows": int(data["event_date"].notna().sum()),
        "distress_keyword_rows_in_window": int(
            (data["distress_keyword_event"] & data["event_date"].between(pd.Timestamp("2010-01-01"), pd.Timestamp("2024-12-31"))).sum()
        ),
        "direct_id_overlap_with_model": int(len(direct_overlap)),
        "conservative_name_map_keys": int(len(name_map)),
        "mapped_distress_event_rows": int(len(events)),
        "unique_event_company_ids": int(events["company_id"].nunique()),
    }
    return events, meta


def add_event_windows(panel: pd.DataFrame, events: pd.DataFrame) -> pd.DataFrame:
    panel = panel.copy()
    panel["fy_end_proxy"] = pd.to_datetime(panel["fiscal_year"].astype("Int64").astype(str) + "-12-31", errors="coerce")
    panel["distress_event_12m_ciq"] = 0
    panel["distress_event_24m_ciq"] = 0
    panel["first_distress_event_date_ciq"] = pd.NaT

    event_map = {
        int(company_id): group["event_date"].sort_values().to_numpy()
        for company_id, group in events.groupby("company_id")
    }
    for idx, row in panel[["company_id", "fy_end_proxy"]].iterrows():
        company_id = row["company_id"]
        fy_end = row["fy_end_proxy"]
        if pd.isna(company_id) or pd.isna(fy_end):
            continue
        dates = event_map.get(int(company_id))
        if dates is None or len(dates) == 0:
            continue
        date_series = pd.Series(pd.to_datetime(dates))
        after = date_series[date_series > fy_end]
        if after.empty:
            continue
        first_after = after.iloc[0]
        panel.at[idx, "first_distress_event_date_ciq"] = first_after
        panel.at[idx, "distress_event_12m_ciq"] = int(first_after <= fy_end + pd.Timedelta(days=365))
        panel.at[idx, "distress_event_24m_ciq"] = int(first_after <= fy_end + pd.Timedelta(days=730))
    return panel


def add_altman_candidates(panel: pd.DataFrame) -> pd.DataFrame:
    panel = panel.copy()
    for col in [
        "total_assets",
        "ebit",
        "revenue",
        "current_assets_ciq_usd_000",
        "current_liabilities_ciq_usd_000",
        "total_liabilities_ciq_usd_000",
        "retained_earnings_ciq_usd_000",
        "market_cap_ciq_usd_m",
    ]:
        if col in panel.columns:
            panel[col] = pd.to_numeric(panel[col], errors="coerce")

    positive_assets = panel["total_assets"].where(panel["total_assets"] > 0)
    positive_liabilities = panel["total_liabilities_ciq_usd_000"].where(panel["total_liabilities_ciq_usd_000"] > 0)
    panel["working_capital_ciq_usd_000"] = panel["current_assets_ciq_usd_000"] - panel["current_liabilities_ciq_usd_000"]
    panel["altman_x1_working_capital_assets_ciq"] = panel["working_capital_ciq_usd_000"] / positive_assets
    panel["altman_x2_retained_earnings_assets_ciq"] = panel["retained_earnings_ciq_usd_000"] / positive_assets
    panel["altman_x3_ebit_assets_ciq"] = panel["ebit"] / positive_assets
    panel["altman_x4_market_value_liabilities_ciq"] = (panel["market_cap_ciq_usd_m"] * 1000.0) / positive_liabilities
    panel["altman_x5_sales_assets_ciq"] = panel["revenue"] / positive_assets
    required = [
        "altman_x1_working_capital_assets_ciq",
        "altman_x2_retained_earnings_assets_ciq",
        "altman_x3_ebit_assets_ciq",
        "altman_x4_market_value_liabilities_ciq",
        "altman_x5_sales_assets_ciq",
    ]
    panel["altman_z_ciq_candidate"] = np.where(
        panel[required].notna().all(axis=1),
        1.2 * panel[required[0]]
        + 1.4 * panel[required[1]]
        + 3.3 * panel[required[2]]
        + 0.6 * panel[required[3]]
        + panel[required[4]],
        np.nan,
    )
    panel["altman_distress_zone_ciq_candidate"] = np.where(
        panel["altman_z_ciq_candidate"].notna(),
        (panel["altman_z_ciq_candidate"] < 1.81).astype(int),
        np.nan,
    )
    panel["altman_gray_zone_ciq_candidate"] = np.where(
        panel["altman_z_ciq_candidate"].notna(),
        panel["altman_z_ciq_candidate"].between(1.81, 2.99, inclusive="both").astype(int),
        np.nan,
    )
    panel["altman_safe_zone_ciq_candidate"] = np.where(
        panel["altman_z_ciq_candidate"].notna(),
        (panel["altman_z_ciq_candidate"] > 2.99).astype(int),
        np.nan,
    )
    return panel


def add_timing_flags(panel: pd.DataFrame) -> pd.DataFrame:
    panel = panel.copy()
    panel["forecast_date_parsed"] = pd.to_datetime(panel.get("forecast_date"), errors="coerce")
    panel["as_of_date_parsed"] = pd.to_datetime(panel.get("as_of_date"), errors="coerce")
    panel["analyst_timing_ok"] = np.where(
        panel["forecast_date_parsed"].notna() & panel["as_of_date_parsed"].notna(),
        (panel["forecast_date_parsed"] <= panel["as_of_date_parsed"]).astype(int),
        np.nan,
    )
    panel["event_window_observable_12m"] = (panel["fiscal_year"] <= 2023).astype(int)
    panel["event_window_observable_24m"] = (panel["fiscal_year"] <= 2022).astype(int)
    return panel


def coverage_table(panel: pd.DataFrame, fields: list[str]) -> pd.DataFrame:
    group_cols = ["market", "fiscal_year"]
    rows = []
    for (market, fiscal_year), group in panel.groupby(group_cols, dropna=False):
        row = {"market": market, "fiscal_year": int(fiscal_year), "rows": int(len(group))}
        for field in fields:
            row[f"{field}_nonmissing"] = int(group[field].notna().sum()) if field in group else 0
            row[f"{field}_missing_pct"] = (
                round(float(group[field].isna().mean() * 100), 2) if field in group and len(group) else 100.0
            )
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["market", "fiscal_year"])


def markdown_table(df: pd.DataFrame, cols: list[str]) -> list[str]:
    header = "| " + " | ".join(cols) + " |"
    divider = "| " + " | ".join("---" for _ in cols) + " |"
    rows = [header, divider]
    for record in df[cols].to_dict("records"):
        values = []
        for col in cols:
            value = record[col]
            if pd.isna(value):
                values.append("")
            else:
                values.append(str(value))
        rows.append("| " + " | ".join(values) + " |")
    return rows


def render_audit(
    panel: pd.DataFrame,
    base: pd.DataFrame,
    metas: dict[str, Any],
    coverage: pd.DataFrame,
    out_panel: Path,
) -> str:
    duplicate_firm_years = int(panel.duplicated(["company_id", "fiscal_year"]).sum())
    model_ids = set(base["company_id"].dropna().astype(int).tolist())
    liq_ids = set(metas["liquidity_market_ids"])
    tl_ids = set(metas["total_liabilities_ids"])
    re_ids = set(metas["retained_earnings_ids"])
    status_ids = set(metas["status_ids"])
    event_ids = set(metas["event_ids"])

    key_fields = [
        "current_assets_ciq_usd_000",
        "current_liabilities_ciq_usd_000",
        "market_cap_ciq_usd_m",
        "total_liabilities_ciq_usd_000",
        "retained_earnings_ciq_usd_000",
        "altman_z_ciq_candidate",
        "distress_event_12m_ciq",
    ]
    lines = [
        "# AEL APAC Firm-Year Panel v2 Capital IQ Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Output panel: `{out_panel}`",
        "",
        "## Decision",
        "",
        "Status: PROVISIONAL_V2_PANEL_BUILT_NOT_SAMPLE_FROZEN",
        "",
        "The script built one auditable v2 panel from the baseline firm-year panel and the Capital IQ supplemental raw workbooks. This is not yet the final submission dataset because sample exclusions, winsorization, and model Go/No-Go decisions still need to be frozen after reviewing this audit.",
        "",
        "## Panel Shape",
        "",
        f"- Baseline rows: {len(base):,}",
        f"- v2 rows: {len(panel):,}",
        f"- Unique companies: {panel['company_id'].nunique():,}",
        f"- Fiscal years: {int(panel['fiscal_year'].min())}-{int(panel['fiscal_year'].max())}",
        f"- Duplicate company-year rows: {duplicate_firm_years:,}",
        "",
        "## Raw Source Evidence",
        "",
    ]
    for group_name in ["liquidity_market", "total_liabilities", "retained_earnings"]:
        lines.append(f"### {group_name}")
        for workbook_meta in metas[group_name]["workbooks"]:
            lines.append(
                f"- `{workbook_meta['name']}`: rows={workbook_meta['company_rows']:,}, sha256=`{workbook_meta['sha256']}`"
            )
        duplicate_pairs = dict(metas[group_name]["duplicate_key_params"])
        if duplicate_pairs:
            sample = dict(list(duplicate_pairs.items())[:20])
            lines.append(f"- Duplicate field-key/parameter pairs handled by first-nonmissing rule: `{sample}`")
        lines.append("")

    lines.extend(
        [
            "### status_universe",
            f"- `{metas['status']['workbook']}`: rows={metas['status']['company_rows']:,}, sha256=`{metas['status']['sha256']}`",
            "",
            "### distress_events",
            f"- `{metas['events']['workbook']}`: raw rows={metas['events']['raw_rows']:,}, parseable dates={metas['events']['parseable_event_rows']:,}, distress-keyword rows in 2010-2024={metas['events']['distress_keyword_rows_in_window']:,}, sha256=`{metas['events']['sha256']}`",
            f"- Direct `SPCIQ ID` overlap with baseline `company_id`: {metas['events']['direct_id_overlap_with_model']:,}",
            f"- Conservative unique-name map keys: {metas['events']['conservative_name_map_keys']:,}",
            f"- Mapped distress-event rows after conservative name matching: {metas['events']['mapped_distress_event_rows']:,}",
            f"- Unique mapped event company IDs: {metas['events']['unique_event_company_ids']:,}",
            "- Event matching is therefore a candidate mapping, not a clean direct-ID mapping; a reviewer-facing strict-event label should either disclose this limitation or be rebuilt from an event export that includes the same Entity ID used in the baseline panel.",
            "",
            "## Model-ID Coverage",
            "",
            "| Source block | Export IDs | Overlap with baseline IDs | Missing baseline IDs |",
            "|---|---:|---:|---:|",
            f"| Liquidity/market | {len(liq_ids):,} | {len(model_ids & liq_ids):,} | {len(model_ids - liq_ids):,} |",
            f"| Total liabilities | {len(tl_ids):,} | {len(model_ids & tl_ids):,} | {len(model_ids - tl_ids):,} |",
            f"| Retained earnings | {len(re_ids):,} | {len(model_ids & re_ids):,} | {len(model_ids - re_ids):,} |",
            f"| Broad status universe | {len(status_ids):,} | {len(model_ids & status_ids):,} | {len(model_ids - status_ids):,} |",
            f"| Distress events | {len(event_ids):,} | {len(model_ids & event_ids):,} | {len(model_ids - event_ids):,} |",
            "",
            "## Key Field Coverage by Market-Year",
            "",
        ]
    )
    display_cols = ["market", "fiscal_year", "rows"]
    for field in key_fields:
        display_cols += [f"{field}_nonmissing", f"{field}_missing_pct"]
    lines.extend(markdown_table(coverage, display_cols))
    lines.extend(
        [
            "",
            "## Label Candidate Rates",
            "",
            "| Label | Non-missing rows | Positive rows | Positive rate among non-missing |",
            "|---|---:|---:|---:|",
        ]
    )
    for label in [
        "altman_distress_zone_ciq_candidate",
        "altman_gray_zone_ciq_candidate",
        "altman_safe_zone_ciq_candidate",
        "distress_event_12m_ciq",
        "distress_event_24m_ciq",
    ]:
        nonmissing = panel[label].notna()
        positives = int((panel.loc[nonmissing, label] == 1).sum())
        denom = int(nonmissing.sum())
        rate = positives / denom if denom else np.nan
        lines.append(f"| {label} | {denom:,} | {positives:,} | {rate:.4f} |")

    forecast_after = int((panel["analyst_timing_ok"] == 0).sum())
    lines.extend(
        [
            "",
            "## Timing and Leakage Checks",
            "",
            f"- Forecast-date rows after as-of date: {forecast_after:,}",
            f"- Rows with 12-month event window observable through 2024: {int(panel['event_window_observable_12m'].sum()):,}",
            f"- Rows with 24-month event window observable through 2024: {int(panel['event_window_observable_24m'].sum()):,}",
            "- Event labels are generated only from post-fiscal-year-end Key Developments dates and are kept as outcome candidates, not predictors.",
            "- Market capitalization uses explicit historical date parameters, not `Current` market cap.",
            "",
            "## Remaining Reviewer Risks",
            "",
            "- `SNL_TOTAL_LIAB` remains a candidate total-liabilities field family and should be disclosed/reconciled before a strict Altman claim.",
            "- Altman formulas are candidate labels until unit consistency, industry exclusions, REIT/fund/financial-firm rules, and winsorization are frozen.",
            "- Event labels rely on keyword filtering of Key Developments text; distress versus non-distress delisting still needs manual or rule-based categorization before being used as a main outcome.",
            "- FY2024 cannot support forward 12m/24m event labels from a 2010-2024 event export; FY2023 cannot support 24m labels.",
            "- This audit does not yet apply final winsorization or exclusion rules.",
            "",
            "## Next Gate",
            "",
            "Proceed to sample-freeze and cleaning rules: exclude/flag ineligible industry structures, deduplicate company-years, set final missingness thresholds, apply winsorization, and then rerun descriptive statistics plus main/robustness models.",
            "",
        ]
    )
    return "\n".join(lines)


def build(out_panel: Path, out_audit: Path) -> None:
    base = pd.read_csv(BASE_PANEL, low_memory=False)
    base["company_id"] = pd.to_numeric(base["company_id"], errors="coerce").astype("Int64")
    base["fiscal_year"] = pd.to_numeric(base["fiscal_year"], errors="coerce").astype("Int64")
    base = base.loc[base["company_id"].notna() & base["fiscal_year"].notna()].copy()
    base["company_id"] = base["company_id"].astype(int)
    base["fiscal_year"] = base["fiscal_year"].astype(int)

    liquidity_market, liquidity_meta = build_liquidity_market(LIQUIDITY_MARKET_WORKBOOKS)
    total_liabilities, total_liab_meta = build_yearly_field(
        TOTAL_LIAB_WORKBOOKS, "SNL_TOTAL_LIAB", "total_liabilities_ciq_usd_000", "total_liabilities"
    )
    retained_earnings, retained_meta = build_yearly_field(
        RETAINED_EARNINGS_WORKBOOKS, "IQ_RETAINED_EARNINGS", "retained_earnings_ciq_usd_000", "retained_earnings"
    )
    status, status_meta = build_status_universe(STATUS_WORKBOOK)
    events, events_meta = load_events(EVENT_WORKBOOK, base)

    panel = base.merge(liquidity_market, on=["company_id", "fiscal_year"], how="left")
    panel = panel.merge(total_liabilities, on=["company_id", "fiscal_year"], how="left")
    panel = panel.merge(retained_earnings, on=["company_id", "fiscal_year"], how="left")
    panel = panel.merge(status, on="company_id", how="left")
    panel = add_event_windows(panel, events)
    panel = add_altman_candidates(panel)
    panel = add_timing_flags(panel)

    fields = [
        "current_assets_ciq_usd_000",
        "current_liabilities_ciq_usd_000",
        "market_cap_ciq_usd_m",
        "total_liabilities_ciq_usd_000",
        "retained_earnings_ciq_usd_000",
        "altman_z_ciq_candidate",
        "distress_event_12m_ciq",
    ]
    coverage = coverage_table(panel, fields)
    metas = {
        "liquidity_market": liquidity_meta,
        "total_liabilities": total_liab_meta,
        "retained_earnings": retained_meta,
        "status": status_meta,
        "events": events_meta,
        "liquidity_market_ids": sorted(liquidity_market["company_id"].dropna().astype(int).unique().tolist()),
        "total_liabilities_ids": sorted(total_liabilities["company_id"].dropna().astype(int).unique().tolist()),
        "retained_earnings_ids": sorted(retained_earnings["company_id"].dropna().astype(int).unique().tolist()),
        "status_ids": sorted(status["company_id"].dropna().astype(int).unique().tolist()),
        "event_ids": sorted(events["company_id"].dropna().astype(int).unique().tolist()),
    }

    out_panel.parent.mkdir(parents=True, exist_ok=True)
    out_audit.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(out_panel, index=False)
    out_audit.write_text(render_audit(panel, base, metas, coverage, out_panel), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-panel", type=Path, default=OUT_PANEL)
    parser.add_argument("--out-audit", type=Path, default=OUT_AUDIT)
    args = parser.parse_args()
    build(args.out_panel, args.out_audit)
    print(f"Wrote {args.out_panel}")
    print(f"Wrote {args.out_audit}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
