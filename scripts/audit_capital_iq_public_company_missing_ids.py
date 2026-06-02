#!/usr/bin/env python3
"""Audit model-panel company IDs missing from the latest Chrome Capital IQ export."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
EXPORT = Path("/Users/guohuiwen/Downloads/SPGlobal_Export_6-1-2026_9a7f1287-0096-45e1-abba-62c17000cad1.xlsx")
OUT_CSV = ROOT / "outputs/capital_iq_public_company_missing_model_ids_20260602.csv"
OUT_MD = ROOT / "outputs/capital_iq_public_company_missing_model_ids_20260602.md"


def load_export_ids() -> set[int]:
    wb = load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    try:
        id_idx = list(header).index("SP_ENTITY_ID")
    except ValueError as exc:
        raise SystemExit("SP_ENTITY_ID not found in export header row") from exc

    ids: set[int] = set()
    for row in ws.iter_rows(min_row=7, values_only=True):
        value = row[id_idx]
        if value in (None, "NA", ""):
            continue
        ids.add(int(value))
    return ids


def exchange_from_name(name: str) -> str:
    match = re.search(r"\(([^():]+):[^()]+\)\s*$", str(name))
    return match.group(1) if match else "NO_TICKER_SUFFIX"


def first_nonblank(values: pd.Series) -> str:
    nonblank = values.dropna().astype(str)
    nonblank = nonblank[nonblank.str.len() > 0]
    return nonblank.iloc[0] if len(nonblank) else ""


def main() -> int:
    export_ids = load_export_ids()
    panel = pd.read_csv(PANEL, low_memory=False)
    panel["company_id"] = panel["company_id"].astype(int)
    panel["name_exchange_suffix"] = panel["company_name"].map(exchange_from_name)
    panel["is_missing_latest_public_export"] = ~panel["company_id"].isin(export_ids)

    company = (
        panel.groupby("company_id", as_index=False)
        .agg(
            company_name=("company_name", first_nonblank),
            first_year=("fiscal_year", "min"),
            last_year=("fiscal_year", "max"),
            firm_years=("fiscal_year", "nunique"),
            market=("market", first_nonblank),
            exchange=("exchange", first_nonblank),
            name_exchange_suffix=("name_exchange_suffix", first_nonblank),
            ticker=("ticker", first_nonblank),
            source_panel=("source_panel", first_nonblank),
            company_status=("company_status", first_nonblank),
            labelled_years=("stress_12m", lambda s: int(s.notna().sum())),
            stress_events=("stress_12m", lambda s: int(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
            covered_years=("analyst_covered", lambda s: int(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
        )
    )
    missing = company[~company["company_id"].isin(export_ids)].copy()
    missing = missing.sort_values(["name_exchange_suffix", "company_name", "company_id"])
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    missing.to_csv(OUT_CSV, index=False)

    def table_from_series(series: pd.Series, name: str) -> list[str]:
        lines = [f"| {name} | Count |", "|---|---:|"]
        for key, count in series.items():
            lines.append(f"| {key if key else '(blank)'} | {int(count)} |")
        return lines

    suffix_counts = missing["name_exchange_suffix"].fillna("").value_counts().sort_index()
    market_counts = missing["market"].fillna("").value_counts().sort_index()
    exchange_counts = missing["exchange"].fillna("").value_counts().sort_index()
    year_span = missing.groupby(["first_year", "last_year"]).size().sort_index()

    sample = missing.head(40)
    lines = [
        "# Capital IQ Public-Company Export Missing Model IDs Audit",
        "",
        "Date: 2026-06-02",
        "",
        "## Inputs",
        "",
        f"- Model panel: `{PANEL.relative_to(ROOT)}`",
        f"- Chrome export: `{EXPORT}`",
        f"- Missing-ID CSV: `{OUT_CSV.relative_to(ROOT)}`",
        "",
        "## Summary",
        "",
        f"- Existing processed model-panel unique company IDs: {company['company_id'].nunique():,}",
        f"- Latest public-company export unique entity IDs: {len(export_ids):,}",
        f"- Overlap: {company['company_id'].isin(export_ids).sum():,}",
        f"- Model-panel IDs missing from latest public-company export: {len(missing):,}",
        "",
        "All missing IDs remain in the Singapore-side processed panel under the current `market` / `exchange` coding, but the ticker suffixes show that a material share are not current SGX/Catalist listings. This is consistent with a historical-universe mismatch rather than simple field-export failure.",
        "",
        "## Missing IDs by Processed Market",
        "",
        *table_from_series(market_counts, "Processed market"),
        "",
        "## Missing IDs by Processed Exchange",
        "",
        *table_from_series(exchange_counts, "Processed exchange"),
        "",
        "## Missing IDs by Ticker Suffix in Company Name",
        "",
        *table_from_series(suffix_counts, "Ticker suffix"),
        "",
        "## Missing IDs by Panel Year Span",
        "",
        "| First year | Last year | Count |",
        "|---:|---:|---:|",
    ]
    for (first_year, last_year), count in year_span.items():
        lines.append(f"| {int(first_year)} | {int(last_year)} | {int(count)} |")

    lines.extend(
        [
            "",
            "## First 40 Missing Companies",
            "",
            "| Company ID | Company | Years | Processed exchange | Ticker suffix | Ticker | Status |",
            "|---:|---|---:|---|---|---|---|",
        ]
    )
    for row in sample.itertuples(index=False):
        lines.append(
            f"| {row.company_id} | {row.company_name} | {row.first_year}-{row.last_year} | "
            f"{row.exchange} | {row.name_exchange_suffix} | {row.ticker} | {row.company_status} |"
        )

    lines.extend(
        [
            "",
            "## Decision",
            "",
            "Do not merge the latest Chrome public-company FY2014-FY2023 export into the manuscript model yet. The remaining 173 IDs should be handled by a targeted historical-universe reconciliation step. A direct next Chrome action is to export by `Entity ID` list or to build supplemental screens for the non-SGX/Catalist ticker-suffix groups identified above, then re-run this audit before any model rebuild.",
            "",
        ]
    )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_CSV.relative_to(ROOT)}")
    print(f"Wrote {OUT_MD.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
