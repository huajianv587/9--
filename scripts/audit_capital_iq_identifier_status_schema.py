#!/usr/bin/env python3
"""Audit Capital IQ identifier/status schema-proof exports."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


NA_STRINGS = {"", "NA", "N/A", "NM", "NONE", "#N/A", "#VALUE!", "#DIV/0!"}
REQUIRED_KEYS = {"SP_ENTITY_ID", "SP_ENTITY_NAME", "SP_EXCHANGE", "SP_COMPANY_TYPE"}
STATUS_KEYS = {"SP_COMPANY_STATUS"}
IPO_KEYS = {"SP_IPO_DATE"}
GEOGRAPHY_KEYS = {"SP_GEOGRAPHY"}
DIRECT_EVENT_KEY_PATTERNS = {
    "status_date": re.compile(r"(STATUS.*DATE|DATE.*STATUS|INACTIVE.*DATE)", re.I),
    "delisting_date": re.compile(r"DELIST.*DATE", re.I),
    "delisting_reason": re.compile(r"DELIST.*REASON|REASON.*DELIST", re.I),
    "bankruptcy_date": re.compile(r"BANKRUPT.*DATE|FILING.*DATE", re.I),
    "liquidation_date": re.compile(r"LIQUIDAT.*DATE", re.I),
    "suspension_date": re.compile(r"SUSPENSION.*DATE|SUSPEND.*DATE", re.I),
}


def clean(value: Any) -> str:
    return str(value or "").strip()


def is_nonmissing(value: Any) -> bool:
    return value is not None and not (
        isinstance(value, str) and value.strip().upper() in NA_STRINGS
    )


def load_workbook_rows(workbook: Path) -> tuple[list[str], list[str], list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError(f"`Sheet1` not found in {workbook}")
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Workbook has too few rows: {workbook}")
    headers = [clean(value) for value in rows[2]]
    keys = [clean(value) for value in rows[3]]
    params = [clean(value) for value in rows[4]]
    data = [row for row in rows[5:] if len(row) > 1 and row[1] not in (None, "")]
    return headers, keys, params, data


def column_indexes(keys: list[str], target_keys: set[str]) -> list[int]:
    return [idx for idx, key in enumerate(keys) if key in target_keys]


def universe_summary(keys: list[str], rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    id_idx = keys.index("SP_ENTITY_ID") if "SP_ENTITY_ID" in keys else 1
    exchange_idx = keys.index("SP_EXCHANGE") if "SP_EXCHANGE" in keys else 2
    type_idx = keys.index("SP_COMPANY_TYPE") if "SP_COMPANY_TYPE" in keys else 3
    ids = []
    exchanges = []
    company_types = []
    for row in rows:
        try:
            ids.append(int(row[id_idx]))
        except (IndexError, TypeError, ValueError):
            pass
        if exchange_idx < len(row) and is_nonmissing(row[exchange_idx]):
            exchanges.append(str(row[exchange_idx]).strip())
        if type_idx < len(row) and is_nonmissing(row[type_idx]):
            company_types.append(str(row[type_idx]).strip())
    return {
        "data_rows": len(rows),
        "unique_ids": len(set(ids)),
        "duplicate_id_rows": len(ids) - len(set(ids)),
        "exchange_counts": Counter(exchanges),
        "company_type_counts": Counter(company_types),
    }


def value_coverage(keys: list[str], rows: list[tuple[Any, ...]], target_keys: set[str]) -> dict[str, Any]:
    coverage = {}
    for idx in column_indexes(keys, target_keys):
        key = keys[idx]
        values = [row[idx] if idx < len(row) else None for row in rows]
        coverage[key] = {
            "column_number": idx + 1,
            "nonmissing": sum(is_nonmissing(value) for value in values),
            "top_values": Counter(str(value).strip() for value in values if is_nonmissing(value)).most_common(20),
        }
    return coverage


def direct_event_key_hits(keys: list[str]) -> dict[str, list[str]]:
    hits: dict[str, list[str]] = {}
    for label, pattern in DIRECT_EVENT_KEY_PATTERNS.items():
        matches = sorted({key for key in keys if key and pattern.search(key)})
        hits[label] = matches
    return hits


def render_report(
    workbook: Path,
    headers: list[str],
    keys: list[str],
    rows: list[tuple[Any, ...]],
    expected_exchanges: list[str],
) -> str:
    universe = universe_summary(keys, rows)
    status_coverage = value_coverage(keys, rows, STATUS_KEYS)
    ipo_coverage = value_coverage(keys, rows, IPO_KEYS)
    geography_coverage = value_coverage(keys, rows, GEOGRAPHY_KEYS)
    direct_hits = direct_event_key_hits(keys)
    detected_required = sorted(REQUIRED_KEYS & set(keys))
    missing_required = sorted(REQUIRED_KEYS - set(keys))

    lines = [
        "# Capital IQ Identifier/Status Schema Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: `{workbook}`",
        "",
        "## Universe",
        "",
        f"- Data rows with Entity ID: {universe['data_rows']}",
        f"- Unique Entity IDs: {universe['unique_ids']}",
        f"- Duplicate Entity-ID rows: {universe['duplicate_id_rows']}",
        f"- Exchange counts: {dict(universe['exchange_counts'])}",
        f"- Company-type counts: {dict(universe['company_type_counts'])}",
        "",
        "## Identifier And Status Keys",
        "",
        f"- Required identifier keys detected: {detected_required}",
        f"- Missing required identifier keys: {missing_required}",
        f"- Company-status coverage: {status_coverage or 'None'}",
        f"- IPO-date coverage: {ipo_coverage or 'None'}",
        f"- Geography coverage: {geography_coverage or 'None'}",
        "",
        "## Direct Event/Boundary Field Search",
        "",
    ]
    for label, matches in direct_hits.items():
        lines.append(f"- {label}: {matches or 'Not detected'}")

    hard_failures = []
    warnings = []
    if missing_required:
        hard_failures.append(f"Missing required identifier keys: {missing_required}.")
    if "SP_COMPANY_STATUS" not in keys:
        hard_failures.append("SP_COMPANY_STATUS not detected.")
    if expected_exchanges:
        exchange_counts = universe["exchange_counts"]
        unexpected = sorted(set(exchange_counts) - set(expected_exchanges))
        missing_expected = [
            exchange for exchange in expected_exchanges if exchange_counts.get(exchange, 0) == 0
        ]
        if missing_expected:
            hard_failures.append(f"Expected exchange(s) not detected: {missing_expected}.")
        if unexpected:
            hard_failures.append(f"Unexpected exchange(s) detected: {unexpected}.")
    if "SP_IPO_DATE" not in keys:
        warnings.append("SP_IPO_DATE not detected; IPO/sample-entry timing remains unavailable in this workbook.")
    missing_direct = [label for label, matches in direct_hits.items() if not matches]
    if missing_direct:
        warnings.append(
            "Direct status/date/reason fields not found for: "
            f"{missing_direct}. Use audited Key Developments raw events for dated distress/delisting evidence."
        )
    if not any(universe["exchange_counts"].values()):
        hard_failures.append("No exchange values detected.")

    lines.extend(["", "## Header Tail", ""])
    tail = [
        (idx + 1, header, key)
        for idx, (header, key) in enumerate(zip(headers, keys))
        if key in STATUS_KEYS | IPO_KEYS | GEOGRAPHY_KEYS
        or any(pattern.search(key) for pattern in DIRECT_EVENT_KEY_PATTERNS.values())
    ]
    if tail:
        lines.append("| Column | Header | Key |")
        lines.append("|---:|---|---|")
        for col, header, key in tail:
            lines.append(f"| {col} | `{header}` | `{key}` |")
    else:
        lines.append("- No status/event-like key columns detected beyond identifier fields.")

    lines.extend(["", "## Audit Decision", ""])
    if hard_failures:
        lines.append("Status: FAIL")
        lines.extend([f"- {item}" for item in hard_failures])
    else:
        lines.append("Status: PASS_WITH_NOTES")
        lines.append("- Company identifier/status schema is present and auditable.")
        lines.extend([f"- {item}" for item in warnings])
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--expected-exchange",
        help="Expected exchange name or comma-separated exchange names, e.g. SGX,Catalist.",
    )
    parser.add_argument("--out-md", type=Path)
    args = parser.parse_args()

    headers, keys, _params, rows = load_workbook_rows(args.workbook)
    expected_exchanges = [
        item.strip()
        for item in (args.expected_exchange or "").split(",")
        if item.strip()
    ]
    report = render_report(args.workbook, headers, keys, rows, expected_exchanges)
    if args.out_md:
        args.out_md.parent.mkdir(parents=True, exist_ok=True)
        args.out_md.write_text(report, encoding="utf-8")
        print(f"Wrote {args.out_md}")
    else:
        print(report)


if __name__ == "__main__":
    main()
