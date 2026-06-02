#!/usr/bin/env python3
"""Audit Capital IQ Key Developments event exports."""

from __future__ import annotations

import argparse
import hashlib
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PANEL = ROOT / "data/processed/ael_apac_firm_year_panel.csv"
START_DATE = pd.Timestamp("2010-01-01")
END_DATE = pd.Timestamp("2024-12-31")

ID_ALIASES = {
    "SP_ENTITY_ID",
    "Entity ID",
    "Company ID",
    "Capital IQ Company ID",
    "IQ ID",
    "SPCIQ ID",
    "company_id",
}
DATE_TOKENS = (
    "key development date",
    "event date",
    "announcement date",
    "announced date",
    "effective date",
    "filing date",
    "date",
)
TYPE_TOKENS = ("key development type", "event type", "development type", "type")
CATEGORY_TOKENS = ("category", "key development category", "event category")
TEXT_TOKENS = ("headline", "title", "description", "summary", "synopsis", "situation", "notes")
DISTRESS_KEYWORDS = (
    "bankrupt",
    "bankruptcy",
    "delist",
    "delisting",
    "suspend",
    "suspension",
    "default",
    "liquidat",
    "receiver",
    "receivership",
    "administration",
    "restructur",
    "distress",
    "going concern",
    "red flag",
    "winding up",
    "insolv",
    "voluntary administration",
)


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_header(labels: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, value in enumerate(labels, start=1):
        label = norm(value) or f"unnamed_{i}"
        if label in seen:
            seen[label] += 1
            label = f"{label}.{seen[label]}"
        else:
            seen[label] = 0
        out.append(label)
    return out


def find_header_row(rows: list[list[Any]]) -> int:
    best_row = 1
    best_score = -1
    tokens = ID_ALIASES | set(DATE_TOKENS) | set(TYPE_TOKENS) | set(CATEGORY_TOKENS) | set(TEXT_TOKENS)
    for idx, row in enumerate(rows, start=1):
        labels = [norm(value) for value in row]
        nonblank = [label for label in labels if label]
        if len(nonblank) < 2:
            continue
        lower = " | ".join(label.lower() for label in nonblank)
        score = min(len(nonblank), 10)
        score += 12 if any(label in ID_ALIASES for label in nonblank) else 0
        score += sum(token.lower() in lower for token in tokens)
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def first_matching(columns: list[str], tokens: tuple[str, ...]) -> str:
    for col in columns:
        lower = col.lower()
        if any(token in lower for token in tokens):
            return col
    return ""


def all_matching(columns: list[str], tokens: tuple[str, ...]) -> list[str]:
    matches = []
    for col in columns:
        lower = col.lower()
        if any(token in lower for token in tokens):
            matches.append(col)
    return matches


def load_sheet(path: Path, sheet_name: str) -> tuple[pd.DataFrame, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    preview = [list(row) for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True)]
    header_row = find_header_row(preview)
    rows = ws.iter_rows(min_row=header_row, values_only=True)
    header = clean_header(list(next(rows)))
    data = pd.DataFrame(rows, columns=header)
    data = data.dropna(how="all")
    if not data.empty:
        code_like = data.astype(str).apply(
            lambda row: sum(value.strip().startswith("SPKD_") for value in row),
            axis=1,
        )
        data = data.loc[code_like < 2]
    return data, header_row


def parse_dates(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    if parsed.notna().sum() == 0:
        parsed = pd.to_datetime(series.astype(str).str.strip(), errors="coerce", dayfirst=False)
    return parsed


def audit_sheet(path: Path, sheet_name: str, model_ids: set[int]) -> dict[str, Any]:
    data, header_row = load_sheet(path, sheet_name)
    columns = list(data.columns)
    id_col = next((col for col in columns if col in ID_ALIASES), "")
    if not id_col:
        id_col = next((col for col in columns if col.lower() in {"sp_entity_id", "entity id", "company id", "company_id"}), "")
    date_cols = all_matching(columns, DATE_TOKENS)
    type_cols = all_matching(columns, TYPE_TOKENS)
    category_cols = all_matching(columns, CATEGORY_TOKENS)
    text_cols = all_matching(columns, TEXT_TOKENS)

    ids = pd.Series(dtype="Int64")
    export_ids: set[int] = set()
    if id_col:
        id_text = data[id_col].astype(str)
        if id_col == "SPCIQ ID":
            id_text = id_text.str.extract(r"(\d+)", expand=False)
        ids = pd.to_numeric(id_text, errors="coerce").dropna().astype(int)
        export_ids = set(ids.tolist())

    parsed_date = pd.Series(pd.NaT, index=data.index, dtype="datetime64[ns]")
    chosen_date_col = date_cols[0] if date_cols else ""
    if chosen_date_col:
        parsed_date = parse_dates(data[chosen_date_col])

    search_cols = list(dict.fromkeys(type_cols + category_cols + text_cols))
    if search_cols:
        text_blob = data[search_cols].fillna("").astype(str).agg(" | ".join, axis=1).str.lower()
    else:
        text_blob = pd.Series("", index=data.index, dtype=str)
    distress_hits = text_blob.apply(lambda value: any(keyword in value for keyword in DISTRESS_KEYWORDS))

    top_types: Counter[str] = Counter()
    for col in type_cols + category_cols:
        top_types.update(data[col].dropna().astype(str).str.strip().replace("", pd.NA).dropna().tolist())

    parsed_count = int(parsed_date.notna().sum())
    in_range = parsed_date.between(START_DATE, END_DATE, inclusive="both")
    out_of_range = parsed_date.notna() & ~in_range

    failures: list[str] = []
    notes: list[str] = []
    if data.empty:
        failures.append("No data rows detected.")
    if not id_col:
        failures.append("No company/entity ID column detected.")
    if not date_cols:
        failures.append("No event/date column detected.")
    if not (type_cols or category_cols or text_cols):
        failures.append("No event type/category/headline/description columns detected.")
    if chosen_date_col and parsed_count == 0:
        failures.append("Event date column exists but no parseable dates were detected.")
    if chosen_date_col and parsed_count > 0 and int(in_range.sum()) == 0:
        failures.append("No parsed event dates fall inside 2010-01-01 to 2024-12-31.")
    if chosen_date_col and int(out_of_range.sum()) > 0:
        notes.append("Some parsed dates fall outside the target 2010-2024 window; review before merge.")
    if int(distress_hits.sum()) == 0:
        notes.append("No distress-keyword hits detected in event text/type fields; this may be an overly broad or wrong export.")

    return {
        "sheet": sheet_name,
        "header_row": header_row,
        "data_rows": int(len(data)),
        "columns": len(columns),
        "id_col": id_col,
        "unique_ids": len(export_ids) if export_ids else 0,
        "model_overlap": len(export_ids & model_ids) if export_ids else 0,
        "extra_ids": len(export_ids - model_ids) if export_ids else 0,
        "missing_model_ids": len(model_ids - export_ids) if export_ids else len(model_ids),
        "date_cols": date_cols[:20],
        "chosen_date_col": chosen_date_col,
        "parsed_dates": parsed_count,
        "date_min": parsed_date.min(),
        "date_max": parsed_date.max(),
        "in_range_dates": int(in_range.sum()),
        "out_of_range_dates": int(out_of_range.sum()),
        "type_cols": type_cols[:20],
        "category_cols": category_cols[:20],
        "text_cols": text_cols[:30],
        "distress_keyword_rows": int(distress_hits.sum()),
        "top_event_labels": top_types.most_common(25),
        "failures": failures,
        "notes": notes,
    }


def fmt_date(value: Any) -> str:
    if pd.isna(value):
        return ""
    return pd.Timestamp(value).date().isoformat()


def render_report(path: Path, audits: list[dict[str, Any]], skipped_sheets: list[str]) -> str:
    stat = path.stat()
    hard_failures = [failure for audit in audits for failure in audit["failures"]]
    status = "FAIL" if hard_failures else "PASS_WITH_NOTES"
    lines = [
        "# Capital IQ Key Developments Event Workbook Audit",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: `{path}`",
        f"Size bytes: {stat.st_size:,}",
        f"Modified: {datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds')}",
        f"SHA-256: `{sha256(path)}`",
        f"Target date window: {START_DATE.date().isoformat()} to {END_DATE.date().isoformat()}",
        "",
        "## Workbook Summary",
        "",
    ]
    if skipped_sheets:
        lines.extend(
            [
                "Skipped metadata sheets: " + ", ".join(f"`{name}`" for name in skipped_sheets),
                "",
            ]
        )
    lines.extend(
        [
            "| Sheet | Rows | Columns | Header row | ID column | Unique IDs | Model overlap | Date column | Parsed dates | In-range dates | Distress-keyword rows |",
            "|---|---:|---:|---:|---|---:|---:|---|---:|---:|---:|",
        ]
    )
    for audit in audits:
        lines.append(
            f"| {audit['sheet']} | {audit['data_rows']} | {audit['columns']} | {audit['header_row']} | "
            f"{audit['id_col']} | {audit['unique_ids']} | {audit['model_overlap']} | "
            f"{audit['chosen_date_col']} | {audit['parsed_dates']} | {audit['in_range_dates']} | "
            f"{audit['distress_keyword_rows']} |"
        )
    for audit in audits:
        lines.extend(
            [
                "",
                f"## Sheet: {audit['sheet']}",
                "",
                f"- Date range parsed: {fmt_date(audit['date_min'])} to {fmt_date(audit['date_max'])}",
                f"- Out-of-range parsed dates: {audit['out_of_range_dates']}",
                f"- Extra export IDs outside model panel: {audit['extra_ids']}",
                f"- Missing model IDs relative to current panel: {audit['missing_model_ids']}",
                "",
                "### Detected columns",
                "",
                f"- Date columns: {audit['date_cols'] or 'None'}",
                f"- Type columns: {audit['type_cols'] or 'None'}",
                f"- Category columns: {audit['category_cols'] or 'None'}",
                f"- Text columns: {audit['text_cols'] or 'None'}",
                "",
                "### Top event labels",
                "",
            ]
        )
        if audit["top_event_labels"]:
            lines.append("| Label | Count |")
            lines.append("|---|---:|")
            for label, count in audit["top_event_labels"]:
                safe_label = str(label).replace("\n", " / ")
                lines.append(f"| {safe_label} | {count} |")
        else:
            lines.append("- None detected")
        lines.extend(["", "### Sheet decision notes", ""])
        if audit["failures"]:
            lines.extend([f"- FAIL: {item}" for item in audit["failures"]])
        if audit["notes"]:
            lines.extend([f"- NOTE: {item}" for item in audit["notes"]])
        if not audit["failures"] and not audit["notes"]:
            lines.append("- Basic event-field and date-window checks passed.")

    lines.extend(["", "## Audit Decision", "", f"Status: {status}"])
    if hard_failures:
        lines.extend([f"- {item}" for item in hard_failures])
    else:
        lines.append("- Event workbook has company IDs, date fields, event descriptors, and parseable dates.")
        lines.append("- Treat this as raw event evidence only; outcome-label construction still needs a separate cleaning script.")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-md", type=Path, required=True)
    args = parser.parse_args()

    workbook = args.workbook.expanduser().resolve()
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")
    panel = pd.read_csv(PANEL, usecols=["company_id"], low_memory=False)
    model_ids = set(pd.to_numeric(panel["company_id"], errors="coerce").dropna().astype(int))
    wb = load_workbook(workbook, read_only=True, data_only=True)
    skipped_sheets = [
        sheet.title
        for sheet in wb.worksheets
        if "criteria" in sheet.title.lower()
    ]
    event_sheets = [
        sheet.title
        for sheet in wb.worksheets
        if sheet.title not in skipped_sheets
    ]
    if not event_sheets:
        raise SystemExit("No event data sheets detected after excluding metadata sheets.")
    audits = [audit_sheet(workbook, sheet_name, model_ids) for sheet_name in event_sheets]
    out = args.out_md if args.out_md.is_absolute() else ROOT / args.out_md
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(render_report(workbook, audits, skipped_sheets), encoding="utf-8")
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
